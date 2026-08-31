from __future__ import annotations

import argparse
import csv
import hashlib
import importlib.util
import json
import math
import os
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any, Iterable


SCRIPT_DIR = Path(__file__).resolve().parent
WORKSPACE_ROOT = SCRIPT_DIR.parents[2]
DEFAULT_CONFIG = SCRIPT_DIR / "config" / "default_config.json"

# VSCode interactive defaults.
# Edit these values, then click "Run Python File" in VSCode. When the script is
# launched without command-line arguments, these settings are used.
INTERACTIVE_ACTION = "build"  # check-env | list-variables | inspect-points | audit-anomalies | build
INTERACTIVE_MODEL = "子模型1"
INTERACTIVE_ALL_MODELS = False
INTERACTIVE_IO_TYPE = "输出"
INTERACTIVE_POINT_CODES = ""
INTERACTIVE_WRITE_REPORT = True
INTERACTIVE_START = ""
INTERACTIVE_END = ""
INTERACTIVE_FREQ_SECONDS = 10
INTERACTIVE_FORCE_REBUILD_CACHES = False
INTERACTIVE_LIMIT_POINTS = 0
INTERACTIVE_CACHES_ONLY = False
INTERACTIVE_CHUNKED = True
INTERACTIVE_CHUNK_DAYS = 1
INTERACTIVE_WRITE_CHUNK_WIDE = False
INTERACTIVE_SEQ_LEN = 120
INTERACTIVE_PRED_LEN = 60
INTERACTIVE_WINDOW_STRIDE = 1
INTERACTIVE_AUDIT_POINT_CODES = "B3TE07D,B3TE19A,B3ZI86,B3TE05,B3_ST19"
INTERACTIVE_AUDIT_FREQ_SECONDS = 300

FALSE_TEXT = {"0", "false", "no", "n", "否", "不纳入", "排除", "剔除"}
TRUE_TEXT = {"1", "true", "yes", "y", "是", "纳入"}


@dataclass(frozen=True)
class VariableSpec:
    model: str
    io_type: str
    role_group: str
    variable_role: str
    point_code: str
    point_name: str
    unit: str
    process_stage: str
    target_name: str
    include_advice: str
    control_priority: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def is_input_role(value: Any) -> bool:
    text = clean_text(value).lower()
    return text in {"输入", "input", "x"} or text.endswith("入")


def is_output_role(value: Any) -> bool:
    text = clean_text(value).lower()
    return text in {"输出", "output", "target", "y"} or text.endswith("出")


def cfg_value(cfg: dict[str, Any], key: str, default: Any, model: str | None = None) -> Any:
    overrides = cfg.get("model_overrides", {})
    if model and isinstance(overrides, dict):
        for name in (model, model_slug(model)):
            values = overrides.get(name)
            if isinstance(values, dict) and key in values:
                return values[key]
    return cfg.get(key, default)


def rel_path(path: Path) -> str:
    try:
        return str(path.resolve().relative_to(WORKSPACE_ROOT.resolve()))
    except ValueError:
        return str(path.resolve())


def resolve_path(raw: str | Path, base: Path = SCRIPT_DIR) -> Path:
    path = Path(raw)
    if path.is_absolute():
        return path
    return (base / path).resolve()


def load_config(path: Path) -> dict[str, Any]:
    with path.open("r", encoding="utf-8") as f:
        cfg = json.load(f)
    cfg["_config_path"] = str(path.resolve())
    cfg["classification_xlsx"] = os.environ.get(
        "WISBURN_FORECAST_CLASSIFICATION_XLSX",
        str(resolve_path(cfg["classification_xlsx"], path.parent)),
    )
    cfg["point_root"] = os.environ.get(
        "WISBURN_FORECAST_POINT_ROOT",
        str(resolve_path(cfg["point_root"], path.parent)),
    )
    artifact_root = os.environ.get("WISBURN_FORECAST_ARTIFACT_ROOT")
    cfg["output_dir"] = str(
        Path(artifact_root) / "data_engineering" / "outputs"
        if artifact_root
        else resolve_path(cfg.get("output_dir", "outputs"), SCRIPT_DIR)
    )
    return cfg


def import_or_none(module_name: str):
    if importlib.util.find_spec(module_name) is None:
        return None
    return __import__(module_name)


def maybe_add_project_site_packages() -> None:
    """Allow running from a normal Python when the project venv is usable."""
    if sys.version_info[:2] != (3, 13):
        return
    site_packages = WORKSPACE_ROOT / ".venv_data_store" / "Lib" / "site-packages"
    if site_packages.exists() and str(site_packages) not in sys.path:
        sys.path.insert(0, str(site_packages))


def require_modules() -> tuple[Any, Any, Any, Any]:
    maybe_add_project_site_packages()
    missing = []
    modules = {}
    for name in ("duckdb", "pandas", "openpyxl", "pyarrow"):
        try:
            modules[name] = __import__(name)
        except Exception:
            missing.append(name)
    if missing:
        raise SystemExit(
            "缺少数据工程依赖: "
            + ", ".join(missing)
            + "\n请在可用的数据工程 Python 环境运行，或执行: python -m pip install duckdb pyarrow pandas openpyxl"
        )
    return modules["duckdb"], modules["pandas"], modules["openpyxl"], modules["pyarrow"]


def check_env(args: argparse.Namespace) -> int:
    print(f"Python: {sys.executable}")
    print(f"Version: {sys.version.split()[0]}")
    maybe_add_project_site_packages()
    ok = True
    for name in ("duckdb", "pandas", "openpyxl", "pyarrow"):
        try:
            module = __import__(name)
            version = getattr(module, "__version__", "unknown")
            print(f"[ok] {name} {version}")
        except Exception as exc:
            ok = False
            print(f"[missing] {name}: {exc}")
    cfg = load_config(args.config)
    for label, path_text in (
        ("classification_xlsx", cfg["classification_xlsx"]),
        ("point_root", cfg["point_root"]),
    ):
        path = Path(path_text)
        print(f"{label}: {rel_path(path)} ({'exists' if path.exists() else 'missing'})")
        ok = ok and path.exists()
    return 0 if ok else 1


def is_false_text(value: str) -> bool:
    return clean_text(value).lower() in FALSE_TEXT or clean_text(value) in FALSE_TEXT


def is_explicitly_excluded(row: dict[str, Any]) -> bool:
    manual = clean_text(row.get("人工纳入"))
    if manual:
        return is_false_text(manual)
    advice = clean_text(row.get("纳入建议"))
    return advice.startswith("不纳入") or "暂不纳入" in advice or "剔除" in advice


def effective_model(row: dict[str, Any]) -> str:
    return clean_text(row.get("人工最终模型")) or clean_text(row.get("建议子模型"))


def effective_role(row: dict[str, Any]) -> str:
    return clean_text(row.get("人工最终角色")) or clean_text(row.get("建议变量角色"))


def load_variable_specs_legacy(cfg: dict[str, Any]) -> list[VariableSpec]:
    maybe_add_project_site_packages()
    try:
        import openpyxl
    except Exception:
        _, _, openpyxl, _ = require_modules()

    xlsx = Path(cfg["classification_xlsx"])
    sheet_name = cfg.get("classification_sheet", "模型变量明细")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Excel 中找不到工作表: {sheet_name}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [clean_text(x) for x in next(rows)]
    required = {"建议子模型", "输入/输出", "角色分组", "点位代码", "点位名称"}
    missing = required.difference(headers)
    if missing:
        raise SystemExit(f"模型变量明细缺少列: {', '.join(sorted(missing))}")

    specs: list[VariableSpec] = []
    seen: set[tuple[str, str, str]] = set()
    valid_models = set(cfg.get("models", []))
    for raw in rows:
        row = dict(zip(headers, raw))
        model = effective_model(row)
        io_type = clean_text(row.get("输入/输出"))
        point_code = clean_text(row.get("点位代码"))
        if not model or not io_type or not point_code:
            continue
        if valid_models and model not in valid_models:
            continue
        if is_explicitly_excluded(row):
            continue
        key = (model, io_type, point_code)
        if key in seen:
            continue
        seen.add(key)
        specs.append(
            VariableSpec(
                model=model,
                io_type=io_type,
                role_group=clean_text(row.get("角色分组")),
                variable_role=effective_role(row),
                point_code=point_code,
                point_name=clean_text(row.get("点位名称")),
                unit=clean_text(row.get("单位")),
                process_stage=clean_text(row.get("工艺阶段")),
                target_name=clean_text(row.get("目标名称")),
                include_advice=clean_text(row.get("纳入建议")),
                control_priority=clean_text(row.get("控制优先级")),
            )
        )
    return specs


def _header(headers: list[str], label: str) -> str:
    """Return a compact-table header whose Chinese label is followed by an optional English alias."""
    for value in headers:
        if value == label or value.startswith(f"{label} "):
            return value
    return ""


def _compact_models(value: Any, valid_models: set[str]) -> list[str]:
    aliases = {
        "model1": "\u5b50\u6a21\u578b1",
        "model2": "\u5b50\u6a21\u578b2",
        "model3": "\u5b50\u6a21\u578b3",
    }
    names: list[str] = []
    for item in clean_text(value).replace("\uff0c", ",").replace("\u3001", ",").replace("/", ",").split(","):
        text = item.strip()
        if not text:
            continue
        model = aliases.get(text.lower(), text)
        if not valid_models or model in valid_models:
            names.append(model)
    return names


def load_variable_specs(cfg: dict[str, Any]) -> list[VariableSpec]:
    """Load either the legacy model table or the current compact point classification table."""
    maybe_add_project_site_packages()
    try:
        import openpyxl
    except Exception:
        _, _, openpyxl, _ = require_modules()

    xlsx = Path(cfg["classification_xlsx"])
    sheet_name = cfg.get("classification_sheet", "\u6a21\u578b\u53d8\u91cf\u660e\u7ec6")
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    if sheet_name not in wb.sheetnames:
        raise SystemExit(f"Excel \u4e2d\u627e\u4e0d\u5230\u5de5\u4f5c\u8868: {sheet_name}")
    ws = wb[sheet_name]
    rows = ws.iter_rows(values_only=True)
    headers = [clean_text(value) for value in next(rows)]
    compact_required = {"\u70b9\u4f4d\u4ee3\u7801", "\u70b9\u4f4d\u540d\u79f0", "\u6a21\u578b\u5f52\u5c5e", "\u53d8\u91cf\u89d2\u8272"}
    compact_columns = {label: _header(headers, label) for label in compact_required}
    if not all(compact_columns.values()):
        return load_variable_specs_legacy(cfg)

    optional_columns = {
        label: _header(headers, label)
        for label in ("\u5355\u4f4d", "\u5b50\u7cfb\u7edf", "\u63a7\u5236\u4f18\u5148\u7ea7", "\u7eb3\u5165\u5efa\u8bae", "\u4eba\u5de5\u5907\u6ce8", "\u53ef\u4fe1\u5ea6\u5907\u6ce8")
    }
    valid_models = set(cfg.get("models", []))
    specs: list[VariableSpec] = []
    seen: set[tuple[str, str, str]] = set()
    for raw in rows:
        row = dict(zip(headers, raw))
        point_code = clean_text(row.get(compact_columns["\u70b9\u4f4d\u4ee3\u7801"]))
        variable_role = clean_text(row.get(compact_columns["\u53d8\u91cf\u89d2\u8272"]))
        if not point_code or not variable_role or is_explicitly_excluded(row):
            continue
        shared_model2_observation = "\u6a21\u578b2\u4ec5\u4f5c\u4e3a\u5171\u4eab\u89c2\u6d4b\u8f93\u5165" in " ".join(
            clean_text(row.get(optional_columns[label]))
            for label in ("\u4eba\u5de5\u5907\u6ce8", "\u53ef\u4fe1\u5ea6\u5907\u6ce8")
        )
        for model in _compact_models(row.get(compact_columns["\u6a21\u578b\u5f52\u5c5e"]), valid_models):
            io_type = (
                "\u8f93\u5165"
                if model == "\u5b50\u6a21\u578b2" and shared_model2_observation
                else ("\u8f93\u51fa" if variable_role == "\u76ee\u6807\u53d8\u91cf" else "\u8f93\u5165")
            )
            key = (model, io_type, point_code)
            if key in seen:
                continue
            seen.add(key)
            specs.append(
                VariableSpec(
                    model=model,
                    io_type=io_type,
                    role_group=variable_role,
                    variable_role=variable_role,
                    point_code=point_code,
                    point_name=clean_text(row.get(compact_columns["\u70b9\u4f4d\u540d\u79f0"])),
                    unit=clean_text(row.get(optional_columns["\u5355\u4f4d"])),
                    process_stage=clean_text(row.get(optional_columns["\u5b50\u7cfb\u7edf"])),
                    target_name=point_code if variable_role == "\u76ee\u6807\u53d8\u91cf" else "",
                    include_advice=clean_text(row.get(optional_columns["\u7eb3\u5165\u5efa\u8bae"])) or "\u7eb3\u5165",
                    control_priority=clean_text(row.get(optional_columns["\u63a7\u5236\u4f18\u5148\u7ea7"])),
                )
            )
    return specs


def point_path(point_root: Path, point_code: str) -> Path:
    return point_root / f"point_code={point_code}" / "data.parquet"


def sql_path(path: Path) -> str:
    return str(path.resolve()).replace("\\", "/").replace("'", "''")


def sql_ts(value: str | None) -> str | None:
    if not value:
        return None
    return clean_text(value).replace("T", " ").replace("'", "''")


def params_hash(payload: dict[str, Any]) -> str:
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:16]


def load_manifest(path: Path) -> dict[str, Any]:
    if not path.exists():
        return {"points": {}}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_manifest(path: Path, manifest: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", encoding="utf-8") as f:
        json.dump(manifest, f, ensure_ascii=False, indent=2)
    try:
        tmp.replace(path)
    except PermissionError:
        with path.open("w", encoding="utf-8") as f:
            json.dump(manifest, f, ensure_ascii=False, indent=2)
        try:
            tmp.unlink()
        except OSError:
            pass


def available_point_codes(point_root: Path) -> set[str]:
    if not point_root.exists():
        return set()
    prefix = "point_code="
    return {
        p.name[len(prefix):]
        for p in point_root.iterdir()
        if p.is_dir() and p.name.startswith(prefix) and (p / "data.parquet").exists()
    }


def derived_source_codes(cfg: dict[str, Any]) -> set[str]:
    out: set[str] = set()
    derived = cfg.get("derived_features", {})
    for item in derived.get("primary_air_sections", []):
        out.update(item.get("sources", []))
    for key in ("secondary_air_flow", "recirculation_air_flow"):
        code = derived.get(key)
        if code:
            out.add(code)
    out.update(derived.get("pusher_position_sources", []))
    for item in derived.get("grate_speed_groups", []):
        out.update(item.get("sources", []))
    for item in cfg.get("shutdown_indicators", []):
        if item.get("point_code"):
            out.add(item["point_code"])
    for item in cfg.get("valve_response_pairs", []):
        if item.get("control"):
            out.add(item["control"])
        if item.get("response"):
            out.add(item["response"])
    return out


def point_codes_for_models(specs: list[VariableSpec], models: set[str], cfg: dict[str, Any]) -> list[str]:
    codes: list[str] = []
    seen: set[str] = set()
    for spec in specs:
        if spec.model in models and spec.point_code not in seen:
            seen.add(spec.point_code)
            codes.append(spec.point_code)
    for code in sorted(derived_source_codes(cfg)):
        if code not in seen:
            seen.add(code)
            codes.append(code)
    return codes


def unique_in_order(values: Iterable[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values:
        if value and value not in seen:
            seen.add(value)
            out.append(value)
    return out


def make_output_dirs(cfg: dict[str, Any], freq_seconds: int) -> dict[str, Path]:
    out = Path(cfg["output_dir"])
    paths = {
        "output": out,
        "cache": out / "cache" / f"resampled_{freq_seconds}s",
        "datasets": out / "datasets",
        "reports": out / "reports",
        "checkpoints": out / "checkpoints",
    }
    for path in paths.values():
        path.mkdir(parents=True, exist_ok=True)
    return paths


def range_suffix(start: Any | None, end: Any | None) -> str:
    if not start and not end:
        return ""
    return f"_{time_label(start or 'begin')}_{time_label(end or 'end')}"


def apply_cache_range(paths: dict[str, Path], freq_seconds: int, start: Any | None, end: Any | None) -> dict[str, Path]:
    suffix = range_suffix(start, end)
    if suffix:
        paths["cache"] = paths["output"] / "cache" / f"resampled_{freq_seconds}s{suffix}"
        paths["cache"].mkdir(parents=True, exist_ok=True)
    return paths


def cache_path(paths: dict[str, Path], point_code: str) -> Path:
    safe = point_code.replace("/", "_").replace("\\", "_").replace(":", "_")
    return paths["cache"] / f"{safe}.parquet"


def resample_point(
    con: Any,
    pd: Any,
    source_path: Path,
    point_code: str,
    output_path: Path,
    *,
    freq_seconds: int,
    start: str | None,
    end: str | None,
) -> dict[str, Any]:
    filters = ["value IS NOT NULL", "isfinite(value)"]
    start_ts = sql_ts(start)
    end_ts = sql_ts(end)
    if start_ts:
        filters.append(f"timestamp >= TIMESTAMP '{start_ts}'")
    if end_ts:
        filters.append(f"timestamp < TIMESTAMP '{end_ts}'")
    where_sql = " AND ".join(filters)
    query = f"""
        SELECT
            TIMESTAMP '1970-01-01'
                + CAST(ceil(epoch(timestamp) / {freq_seconds}) * {freq_seconds} AS BIGINT) * INTERVAL '1 second'
                AS timestamp,
            avg(value) AS value,
            count(*) AS raw_count
        FROM read_parquet('{sql_path(source_path)}')
        WHERE {where_sql}
        GROUP BY 1
        ORDER BY 1
    """
    frame = con.execute(query).fetch_df()
    if frame.empty:
        output_path.parent.mkdir(parents=True, exist_ok=True)
        frame = pd.DataFrame({"timestamp": pd.to_datetime([]), point_code: []})
    else:
        frame["timestamp"] = pd.to_datetime(frame["timestamp"])
        frame = frame.rename(columns={"value": point_code})
        frame = frame[["timestamp", point_code, "raw_count"]]
    frame.to_parquet(output_path, index=False)

    non_null = int(frame[point_code].notna().sum()) if point_code in frame else 0
    row_count = int(len(frame))
    if row_count > 1:
        start_time = str(frame["timestamp"].iloc[0])
        end_time = str(frame["timestamp"].iloc[-1])
    elif row_count == 1:
        start_time = end_time = str(frame["timestamp"].iloc[0])
    else:
        start_time = end_time = ""
    return {
        "point_code": point_code,
        "rows": row_count,
        "non_null_rows": non_null,
        "missing_rate": float(1 - non_null / row_count) if row_count else 1.0,
        "start_time": start_time,
        "end_time": end_time,
        "output": str(output_path),
    }


def resample_needed(
    manifest_item: dict[str, Any] | None,
    source_path: Path,
    output_path: Path,
    param_id: str,
    force: bool,
) -> bool:
    if force or not output_path.exists() or manifest_item is None:
        return True
    stat = source_path.stat()
    return not (
        manifest_item.get("status") == "built"
        and manifest_item.get("param_id") == param_id
        and manifest_item.get("source_mtime") == stat.st_mtime
        and manifest_item.get("source_size") == stat.st_size
    )


def build_point_caches(
    cfg: dict[str, Any],
    specs: list[VariableSpec],
    models: set[str],
    *,
    freq_seconds: int,
    start: str | None,
    end: str | None,
    force: bool,
    limit_points: int,
) -> tuple[list[str], list[dict[str, Any]]]:
    duckdb, pd, _, _ = require_modules()
    point_root = Path(cfg["point_root"])
    paths = apply_cache_range(make_output_dirs(cfg, freq_seconds), freq_seconds, start, end)
    manifest_path = paths["checkpoints"] / f"resample_manifest_{freq_seconds}s{range_suffix(start, end)}.json"
    manifest = load_manifest(manifest_path)
    manifest.setdefault("points", {})

    codes = point_codes_for_models(specs, models, cfg)
    if limit_points > 0:
        codes = codes[:limit_points]
    param_id = params_hash({"freq_seconds": freq_seconds, "start": start, "end": end})
    con = duckdb.connect(":memory:")
    quality_rows: list[dict[str, Any]] = []
    built_codes: list[str] = []
    total = len(codes)
    for idx, code in enumerate(codes, 1):
        src = point_path(point_root, code)
        out = cache_path(paths, code)
        if not src.exists():
            print(f"[{idx}/{total}] missing source {code}", flush=True)
            row = {"point_code": code, "status": "missing_source", "source": str(src)}
            manifest["points"][code] = row
            quality_rows.append(row)
            continue
        item = manifest["points"].get(code)
        stat = src.stat()
        if not resample_needed(item, src, out, param_id, force):
            print(f"[{idx}/{total}] skip {code}", flush=True)
            built_codes.append(code)
            quality_rows.append(item)
            continue
        try:
            print(f"[{idx}/{total}] build {code}", flush=True)
            row = resample_point(
                con,
                pd,
                src,
                code,
                out,
                freq_seconds=freq_seconds,
                start=start,
                end=end,
            )
            row.update(
                {
                    "status": "built",
                    "source": str(src),
                    "source_mtime": stat.st_mtime,
                    "source_size": stat.st_size,
                    "param_id": param_id,
                }
            )
            built_codes.append(code)
        except Exception as exc:
            row = {
                "point_code": code,
                "status": "failed",
                "error": str(exc),
                "source": str(src),
                "source_mtime": stat.st_mtime,
                "source_size": stat.st_size,
                "param_id": param_id,
            }
            print(f"[{idx}/{total}] failed {code}: {exc}", flush=True)
        manifest["points"][code] = row
        quality_rows.append(row)
        save_manifest(manifest_path, manifest)

    report_path = paths["reports"] / f"point_quality_{freq_seconds}s.csv"
    pd.DataFrame(quality_rows).to_csv(report_path, index=False, encoding="utf-8-sig")
    print(f"point quality report: {rel_path(report_path)}", flush=True)
    return built_codes, quality_rows


def read_cached_series(
    pd: Any,
    paths: dict[str, Path],
    point_code: str,
    *,
    start: Any | None = None,
    end: Any | None = None,
) -> Any:
    path = cache_path(paths, point_code)
    filters = []
    if start is not None:
        filters.append(("timestamp", ">=", pd.Timestamp(start).to_pydatetime()))
    if end is not None:
        filters.append(("timestamp", "<", pd.Timestamp(end).to_pydatetime()))
    try:
        frame = pd.read_parquet(path, filters=filters or None)
    except Exception:
        frame = pd.read_parquet(path)
        if start is not None:
            frame = frame[frame["timestamp"] >= pd.Timestamp(start)]
        if end is not None:
            frame = frame[frame["timestamp"] < pd.Timestamp(end)]
    if point_code not in frame.columns:
        return None
    return frame[["timestamp", point_code]].drop_duplicates("timestamp", keep="last")


def estimate_wide_table_raw_gb(paths: dict[str, Path], point_codes: Iterable[str]) -> tuple[int, int, float]:
    import pyarrow.parquet as pq

    max_rows = 0
    found_cols = 0
    for code in unique_in_order(point_codes):
        path = cache_path(paths, code)
        if not path.exists():
            continue
        found_cols += 1
        max_rows = max(max_rows, int(pq.ParquetFile(path).metadata.num_rows))
    raw_gb = max_rows * max(found_cols, 1) * 8 / (1024**3)
    return max_rows, found_cols, raw_gb


def check_wide_table_size(paths: dict[str, Path], point_codes: Iterable[str], cfg: dict[str, Any]) -> None:
    max_raw_gb = float(cfg.get("max_wide_table_raw_gb", 6))
    rows, cols, raw_gb = estimate_wide_table_raw_gb(paths, point_codes)
    if rows == 0 or cols == 0 or raw_gb <= max_raw_gb:
        return
    raise SystemExit(
        "\n宽表预计过大，已提前停止，避免 pandas 合并时耗尽内存。\n"
        f"- 当前缓存规模：约 {rows:,} 行 x {cols:,} 点位\n"
        f"- float64 原始矩阵约 {raw_gb:.1f} GB；pandas outer merge 的实际峰值通常会更高\n"
        f"- 当前保护阈值 max_wide_table_raw_gb = {max_raw_gb:.1f} GB\n\n"
        "建议：\n"
        "1. 先设置 INTERACTIVE_START / INTERACTIVE_END，只跑 3 小时或 1 天测试。\n"
        "2. 如果要看全年趋势，把 INTERACTIVE_FREQ_SECONDS 改成 10、60 或 300。\n"
        "3. 如果必须做全年 1 秒训练集，需要改成按天/周分块构建并保留 1500 秒窗口重叠。"
    )


def merge_wide_table(
    pd: Any,
    paths: dict[str, Path],
    point_codes: Iterable[str],
    cfg: dict[str, Any],
    *,
    start: Any | None = None,
    end: Any | None = None,
    check_size: bool = True,
) -> Any:
    point_codes = unique_in_order(point_codes)
    if check_size:
        check_wide_table_size(paths, point_codes, cfg)
    wide = None
    for idx, code in enumerate(point_codes, 1):
        path = cache_path(paths, code)
        if not path.exists():
            continue
        frame = read_cached_series(pd, paths, code, start=start, end=end)
        if frame is None:
            continue
        if wide is None:
            wide = frame
        else:
            wide = wide.merge(frame, on="timestamp", how="outer")
        if idx % 25 == 0:
            print(f"  merged {idx} point caches", flush=True)
    if wide is None:
        return pd.DataFrame({"timestamp": pd.to_datetime([])})
    return wide.sort_values("timestamp", kind="stable").reset_index(drop=True)


def deadband_filter(pd: Any, series: Any, threshold_pct: float = 0.04, min_abs_diff: float | None = None) -> Any:
    data = series.to_numpy(dtype=float, copy=True)
    if len(data) == 0 or bool(pd.isna(data).all()):
        return series
    valid = ~pd.isna(data)
    first = int(valid.argmax())
    current = float(data[first])
    out = data.copy()
    out[: first + 1] = current
    for i in range(first + 1, len(data)):
        if pd.isna(data[i]):
            out[i] = current
            continue
        threshold = max(abs(current) * threshold_pct, 1e-6)
        if min_abs_diff is not None:
            threshold = max(threshold, min_abs_diff)
        if abs(float(data[i]) - current) > threshold:
            current = float(data[i])
        out[i] = current
    return pd.Series(out, index=series.index)


def add_derived_features(pd: Any, df: Any, cfg: dict[str, Any]) -> tuple[Any, list[dict[str, str]]]:
    out = df.copy()
    notes: list[dict[str, str]] = []
    derived = cfg.get("derived_features", {})

    section_cols = []
    for item in derived.get("primary_air_sections", []):
        sources = [c for c in item.get("sources", []) if c in out.columns]
        name = item.get("name")
        if not name:
            continue
        if sources:
            out[name] = out[sources].mean(axis=1)
            out[name] = deadband_filter(pd, out[name], threshold_pct=0.04, min_abs_diff=200)
            section_cols.append(name)
        else:
            notes.append({"feature": name, "status": "missing_all_sources", "sources": ",".join(item.get("sources", []))})

    if section_cols:
        total_name = "primary_air_total_calc"
        out[total_name] = out[section_cols].sum(axis=1, min_count=1)
        denom = out[total_name].replace(0, pd.NA)
        for idx, col in enumerate(section_cols, 1):
            ratio = f"primary_air_section_{idx}_ratio"
            out[ratio] = out[col] / denom
            out[ratio] = deadband_filter(pd, out[ratio], threshold_pct=0.05)

        secondary = derived.get("secondary_air_flow")
        if secondary in out.columns:
            out["primary_air_ratio_primary_secondary"] = out[total_name] / (out[total_name] + out[secondary] + 1e-5)
            out["primary_air_ratio_primary_secondary"] = deadband_filter(
                pd, out["primary_air_ratio_primary_secondary"], threshold_pct=0.05
            )
        else:
            notes.append({"feature": "primary_air_ratio_primary_secondary", "status": "missing_secondary_air", "sources": str(secondary)})

        recirc = derived.get("recirculation_air_flow")
        if recirc and recirc in out.columns and secondary in out.columns:
            total_air = out[total_name] + out[secondary] + out[recirc]
            out["recirculation_air_ratio_total"] = out[recirc] / (total_air + 1e-5)
            out["recirculation_air_ratio_secondary"] = out[recirc] / (out[recirc] + out[secondary] + 1e-5)
        else:
            notes.append({"feature": "recirculation_air_ratios", "status": "not_built_no_recirc_flow", "sources": str(recirc)})

    pusher_sources = [c for c in derived.get("pusher_position_sources", []) if c in out.columns]
    if pusher_sources:
        out["pusher_position_mean"] = out[pusher_sources].mean(axis=1)
        out["pusher_position_mean"] = deadband_filter(pd, out["pusher_position_mean"], threshold_pct=0.04, min_abs_diff=2)

    for item in derived.get("grate_speed_groups", []):
        sources = [c for c in item.get("sources", []) if c in out.columns]
        name = item.get("name")
        if name and sources:
            out[name] = out[sources].mean(axis=1)
            out[name] = deadband_filter(pd, out[name], threshold_pct=0.04, min_abs_diff=2)

    return out, notes


def robust_clean_series(pd: Any, series: Any, *, min_val: float | None = None, max_val: float | None = None) -> Any:
    s = series.copy()
    if min_val is not None:
        s[s < min_val] = pd.NA
    if max_val is not None:
        s[s > max_val] = pd.NA
    if s.isna().all():
        return s
    diff = s.diff().abs()
    diff_med = diff.median(skipna=True)
    if diff_med and not math.isnan(float(diff_med)) and float(diff_med) > 0:
        s[diff > float(diff_med) * 10.0] = pd.NA
    median = s.rolling(window=15, center=False, min_periods=3).median()
    mad = (s - median).abs().rolling(window=15, center=False, min_periods=3).median()
    mask = (s - median).abs() > (3 * 1.4826 * mad + 1e-6)
    s[mask.fillna(False)] = pd.NA
    return s


def range_for_spec(spec: VariableSpec, cfg: dict[str, Any]) -> tuple[float | None, float | None]:
    text = f"{spec.point_code} {spec.point_name} {spec.target_name}".upper()
    for token, bounds in cfg.get("pollutant_ranges", {}).items():
        if token.upper() in text:
            return float(bounds[0]), float(bounds[1])
    if "%" in spec.unit:
        return 0.0, 100.0
    return None, None


def clean_model_frame(pd: Any, df: Any, specs: list[VariableSpec], cfg: dict[str, Any], *, model: str | None = None) -> Any:
    out = df.copy()
    active_model = model or (specs[0].model if specs else None)
    by_code = {s.point_code: s for s in specs}
    targets = [s.point_code for s in specs if is_output_role(s.io_type) and s.point_code in out.columns]
    inputs = [s.point_code for s in specs if is_input_role(s.io_type) and s.point_code in out.columns]

    for code in targets + inputs:
        spec = by_code.get(code)
        if not spec:
            continue
        min_val, max_val = range_for_spec(spec, cfg)
        if min_val is not None or max_val is not None or is_output_role(spec.io_type):
            out[code] = robust_clean_series(pd, out[code], min_val=min_val, max_val=max_val)

    # Targets must remain causal: trailing smoothing and limited ffill only.
    for code in targets:
        out[code] = out[code].rolling(window=3, center=False, min_periods=1).mean()
        out[code] = out[code].ffill(limit=int(cfg_value(cfg, "target_forward_fill_limit", 5, active_model)))

    # Inputs can use short interpolation because they are known covariates/states.
    if inputs:
        out[inputs] = out[inputs].interpolate(
            method="linear",
            limit=int(cfg_value(cfg, "input_interpolate_limit", 6, active_model)),
            limit_direction="both",
        )
        out[inputs] = out[inputs].ffill(limit=int(cfg_value(cfg, "input_forward_fill_limit", 30, active_model)))
    return out


def quality_summary(pd: Any, df: Any, point_codes: list[str], cfg: dict[str, Any]) -> Any:
    rows = []
    zero_abs_tolerance = float(cfg.get("zero_abs_tolerance", 1e-9))
    for code in point_codes:
        if code not in df.columns:
            rows.append({"point_code": code, "status": "missing_column"})
            continue
        s = df[code]
        non_null = int(s.notna().sum())
        zero_ratio = float((s.dropna().abs() <= zero_abs_tolerance).mean()) if non_null else 1.0
        constant_ratio = float(s.value_counts(dropna=True, normalize=True).iloc[0]) if non_null else 1.0
        missing_rate = float(1 - non_null / len(s)) if len(s) else 1.0
        rows.append(
            {
                "point_code": code,
                "rows": int(len(s)),
                "non_null_rows": non_null,
                "missing_rate": missing_rate,
                "zero_ratio": zero_ratio,
                "constant_ratio": constant_ratio,
                "min": float(s.min(skipna=True)) if non_null else math.nan,
                "max": float(s.max(skipna=True)) if non_null else math.nan,
                "mean": float(s.mean(skipna=True)) if non_null else math.nan,
                "warn_missing": bool(missing_rate > float(cfg.get("missing_rate_warn", 0.3))) if len(s) else True,
                "warn_zero": bool(zero_ratio > float(cfg.get("zero_ratio_warn", 0.95))),
                "warn_constant": bool(constant_ratio > float(cfg.get("constant_ratio_warn", 0.95))),
            }
        )
    return pd.DataFrame(rows)


def low_information_inputs(pd: Any, quality: Any, specs: list[VariableSpec], cfg: dict[str, Any]) -> Any:
    input_codes = {s.point_code for s in specs if is_input_role(s.io_type)}
    keep_codes = set(cfg.get("low_information_keep_point_codes", []))
    rows = []
    for _, row in quality.iterrows():
        code = row.get("point_code")
        if code not in input_codes or code in keep_codes:
            continue
        zero_ratio = float(row.get("zero_ratio", 0) or 0)
        constant_ratio = float(row.get("constant_ratio", 0) or 0)
        reasons = []
        if zero_ratio >= float(cfg.get("low_information_zero_ratio_drop", 0.98)):
            reasons.append("near_zero")
        if constant_ratio >= float(cfg.get("low_information_constant_ratio_drop", 0.98)):
            reasons.append("constant")
        if reasons:
            rows.append(
                {
                    "point_code": code,
                    "zero_ratio": zero_ratio,
                    "constant_ratio": constant_ratio,
                    "action": "drop_from_training_inputs",
                    "reason": "|".join(reasons),
                }
            )
    return pd.DataFrame(rows)


def unusable_outputs(pd: Any, quality: Any, specs: list[VariableSpec], cfg: dict[str, Any]) -> Any:
    output_codes = {s.point_code for s in specs if is_output_role(s.io_type)}
    max_missing = float(cfg.get("max_target_missing_rate_drop", 0.98))
    max_constant = float(cfg.get("max_target_constant_ratio_drop", 0.98))
    min_non_null = int(cfg.get("min_target_non_null_rows", 1))
    rows = []
    for _, row in quality.iterrows():
        code = row.get("point_code")
        if code not in output_codes:
            continue
        raw_non_null = row.get("non_null_rows", 0)
        raw_missing_rate = row.get("missing_rate", 1)
        raw_constant_ratio = row.get("constant_ratio", 1)
        non_null = int(float(raw_non_null)) if raw_non_null not in (None, "") else 0
        missing_rate = float(raw_missing_rate) if raw_missing_rate not in (None, "") else 1.0
        constant_ratio = float(raw_constant_ratio) if raw_constant_ratio not in (None, "") else 1.0
        reasons = []
        if non_null < min_non_null:
            reasons.append("too_few_non_null")
        if missing_rate >= max_missing:
            reasons.append("too_many_missing")
        if constant_ratio >= max_constant:
            reasons.append("constant")
        if reasons:
            rows.append(
                {
                    "point_code": code,
                    "non_null_rows": non_null,
                    "missing_rate": missing_rate,
                    "constant_ratio": constant_ratio,
                    "action": "drop_from_training_outputs",
                    "reason": "|".join(reasons),
                }
            )
    return pd.DataFrame(rows)


def usable_output_codes(specs: list[VariableSpec], df: Any, excluded_outputs: Iterable[str]) -> list[str]:
    excluded = set(excluded_outputs)
    return [
        s.point_code
        for s in specs
        if is_output_role(s.io_type) and s.point_code in df.columns and s.point_code not in excluded
    ]


def boolean_windows(
    pd: Any,
    timestamps: Any,
    mask: Any,
    reason: str,
    *,
    max_gap_seconds: float = 120,
    min_duration_seconds: float = 0,
) -> Any:
    temp = pd.DataFrame({"timestamp": timestamps, "flag": mask.fillna(False).to_numpy()})
    temp = temp[temp["flag"]].copy()
    if temp.empty:
        return pd.DataFrame(columns=["start_time", "end_time", "rows", "reason"])
    gap = temp["timestamp"].diff().dt.total_seconds().fillna(0)
    group = (gap > max_gap_seconds).cumsum()
    rows = []
    for _, g in temp.groupby(group):
        start_time = g["timestamp"].iloc[0]
        end_time = g["timestamp"].iloc[-1]
        duration = float((end_time - start_time).total_seconds()) if len(g) > 1 else 0.0
        if duration < min_duration_seconds:
            continue
        rows.append(
            {
                "start_time": start_time,
                "end_time": end_time,
                "rows": len(g),
                "duration_seconds": duration,
                "reason": reason,
            }
        )
    return pd.DataFrame(rows)


def windows_to_mask(pd: Any, timestamps: Any, windows: Any) -> Any:
    mask = pd.Series(False, index=timestamps.index)
    if windows is None or len(windows) == 0:
        return mask
    for _, row in windows.iterrows():
        start_time = pd.to_datetime(row["start_time"])
        end_time = pd.to_datetime(row["end_time"])
        mask |= (timestamps >= start_time) & (timestamps <= end_time)
    return mask


def detect_excluded_windows(pd: Any, df: Any, cfg: dict[str, Any], model_point_codes: list[str]) -> Any:
    masks = []
    reasons = []
    existing = [c for c in model_point_codes if c in df.columns]
    if existing:
        collective_missing = df[existing].isna().mean(axis=1)
        masks.append(collective_missing >= float(cfg.get("collective_missing_rate_warn", 0.4)))
        reasons.append("collective_missing")
    for item in cfg.get("shutdown_indicators", []):
        code = item.get("point_code")
        if code in df.columns:
            low_value = df[code] < float(item.get("min_running_value", 0))
            if bool(item.get("missing_is_bad", False)):
                masks.append(df[code].isna() | low_value)
            else:
                masks.append(df[code].notna() & low_value)
            reasons.append(f"shutdown_indicator:{code}")
    windows = [boolean_windows(pd, df["timestamp"], mask, reason) for mask, reason in zip(masks, reasons)]
    if not windows:
        return pd.DataFrame(columns=["start_time", "end_time", "rows", "reason"])
    return pd.concat(windows, ignore_index=True)


def long_common_low_anomaly(
    pd: Any,
    df: Any,
    cfg: dict[str, Any],
    model_point_codes: list[str],
    *,
    freq_seconds: int,
) -> tuple[Any, Any, Any]:
    settings = cfg.get("long_common_low_anomaly", {})
    empty_windows = pd.DataFrame(columns=["start_time", "end_time", "rows", "duration_seconds", "reason"])
    empty_candidates = pd.DataFrame(
        columns=["point_code", "low_cut", "missing_rate", "zero_ratio", "constant_ratio", "q_low", "q50", "q95"]
    )
    empty_score = pd.DataFrame({"timestamp": df["timestamp"], "common_low_rate": 0.0, "bad_common_long_low": False})
    if not settings or not bool(settings.get("enabled", True)):
        return empty_score, empty_windows, empty_candidates

    zero_abs_tolerance = float(cfg.get("zero_abs_tolerance", 1e-9))
    candidate_rows = []
    low_flags = []
    for code in unique_in_order(model_point_codes):
        if code not in df.columns:
            continue
        s = df[code]
        if not str(s.dtype).startswith(("float", "int")):
            continue
        non_null = int(s.notna().sum())
        if non_null == 0 or len(s) == 0:
            continue
        missing_rate = float(1 - non_null / len(s))
        values = s.dropna()
        zero_ratio = float((values.abs() <= zero_abs_tolerance).mean())
        constant_ratio = float(values.value_counts(dropna=True, normalize=True).iloc[0]) if non_null else 1.0
        if missing_rate > float(settings.get("max_candidate_missing_rate", 0.4)):
            continue
        if zero_ratio > float(settings.get("max_candidate_zero_ratio", 0.98)):
            continue
        if constant_ratio > float(settings.get("max_candidate_constant_ratio", 0.98)):
            continue

        q_low = float(values.quantile(float(settings.get("low_quantile", 0.08))))
        q50 = float(values.quantile(0.50))
        q95 = float(values.quantile(0.95))
        if not (q95 > q_low):
            continue
        low_cut = q_low + max(q50 - q_low, 0.0) * float(settings.get("low_margin_ratio", 0.08))
        low_flags.append((code, (s <= low_cut).fillna(False)))
        candidate_rows.append(
            {
                "point_code": code,
                "low_cut": low_cut,
                "missing_rate": missing_rate,
                "zero_ratio": zero_ratio,
                "constant_ratio": constant_ratio,
                "q_low": q_low,
                "q50": q50,
                "q95": q95,
            }
        )

    candidates = pd.DataFrame(candidate_rows)
    min_cols = int(settings.get("min_candidate_columns", 5))
    if len(low_flags) < min_cols:
        return empty_score, empty_windows, candidates

    low_frame = pd.DataFrame({code: flag.to_numpy() for code, flag in low_flags})
    common_low_rate = low_frame.mean(axis=1)
    rolling_steps = max(1, int(math.ceil(float(settings.get("rolling_seconds", 1800)) / max(freq_seconds, 1))))
    common_low_smoothed = common_low_rate.rolling(rolling_steps, center=True, min_periods=max(1, rolling_steps // 3)).mean()
    raw_bad = common_low_smoothed >= float(settings.get("common_low_rate_threshold", 0.25))

    windows = boolean_windows(
        pd,
        df["timestamp"],
        raw_bad,
        "long_common_low",
        max_gap_seconds=float(settings.get("bridge_gap_seconds", 3600)),
        min_duration_seconds=float(settings.get("min_duration_seconds", 21600)),
    )
    final_bad = windows_to_mask(pd, df["timestamp"], windows)
    score = pd.DataFrame(
        {
            "timestamp": df["timestamp"],
            "common_low_rate": common_low_rate,
            "common_low_rate_smoothed": common_low_smoothed,
            "bad_common_long_low": final_bad,
        }
    )
    return score, windows, candidates


def expand_boolean_mask(pd: Any, mask: Any, buffer_steps: int) -> Any:
    if buffer_steps <= 0 or len(mask) == 0:
        return mask.fillna(False)
    width = buffer_steps * 2 + 1
    expanded = mask.fillna(False).astype(int).rolling(width, center=True, min_periods=1).max()
    return expanded.astype(bool)


def build_training_mask(
    pd: Any,
    df: Any,
    cfg: dict[str, Any],
    model_point_codes: list[str],
    output_codes: list[str],
    *,
    freq_seconds: int,
    common_low_score: Any | None = None,
    model: str | None = None,
) -> Any:
    mask = pd.DataFrame({"timestamp": df["timestamp"]})
    existing = [c for c in model_point_codes if c in df.columns]
    if existing:
        mask["model_valid_rate"] = df[existing].notna().mean(axis=1)
        mask["collective_missing_rate"] = df[existing].isna().mean(axis=1)
    else:
        mask["model_valid_rate"] = 0.0
        mask["collective_missing_rate"] = 1.0

    mask["bad_collective_missing"] = mask["collective_missing_rate"] >= float(cfg.get("collective_missing_rate_warn", 0.4))
    mask["bad_low_valid_rate"] = mask["model_valid_rate"] < float(cfg.get("min_valid_model_columns_rate", 0.65))
    if output_codes:
        mask["bad_target_missing"] = df[output_codes].isna().any(axis=1)
    else:
        mask["bad_target_missing"] = True

    shutdown_cols = []
    for item in cfg.get("shutdown_indicators", []):
        code = item.get("point_code")
        col = f"bad_shutdown_{code}"
        if code in df.columns:
            low_value = df[code] < float(item.get("min_running_value", 0))
            if bool(item.get("missing_is_bad", False)):
                mask[col] = df[code].isna() | low_value
            else:
                mask[col] = df[code].notna() & low_value
        else:
            mask[col] = False
        shutdown_cols.append(col)

    if common_low_score is not None and "bad_common_long_low" in common_low_score.columns:
        mask["bad_common_long_low"] = common_low_score["bad_common_long_low"].fillna(False).to_numpy()
    else:
        mask["bad_common_long_low"] = False

    operational_cols = ["bad_collective_missing", "bad_common_long_low"] + shutdown_cols
    mask["bad_operational_raw"] = mask[operational_cols].any(axis=1) if operational_cols else False
    buffer_seconds = int(cfg_value(cfg, "training_exclusion_buffer_seconds", 0, model))
    buffer_steps = int(math.ceil(buffer_seconds / max(freq_seconds, 1)))
    mask["bad_operational_buffered"] = expand_boolean_mask(pd, mask["bad_operational_raw"], buffer_steps)
    mask["valid_for_training"] = ~(
        mask["bad_operational_buffered"] | mask["bad_low_valid_rate"] | mask["bad_target_missing"]
    )
    return mask


def build_window_index(
    pd: Any,
    df: Any,
    valid_mask: Any,
    *,
    seq_len: int,
    pred_len: int,
    stride: int,
    freq_seconds: int,
) -> Any:
    total_len = int(seq_len) + int(pred_len)
    if len(df) < total_len or total_len <= 0:
        return pd.DataFrame(
            columns=[
                "start_pos",
                "input_start_pos",
                "input_end_pos",
                "target_start_pos",
                "target_end_pos",
                "input_start_time",
                "input_end_time",
                "target_start_time",
                "target_end_time",
            ]
        )

    valid = valid_mask.fillna(False).astype(int)
    valid_forward = valid.rolling(total_len, min_periods=total_len).sum().shift(-(total_len - 1)) == total_len

    step_seconds = df["timestamp"].diff().dt.total_seconds()
    bad_step = step_seconds.ne(float(freq_seconds))
    if len(bad_step):
        bad_step.iloc[0] = False
    bad_forward = bad_step.astype(int).rolling(total_len - 1, min_periods=total_len - 1).sum().shift(-(total_len - 1))
    regular_forward = bad_forward.fillna(1).eq(0)

    stride = max(1, int(stride))
    start_pos = pd.Series(range(len(df)), index=df.index)
    stride_ok = (start_pos % stride) == 0
    ok = valid_forward.fillna(False) & regular_forward & stride_ok
    starts = start_pos[ok].astype(int)
    if starts.empty:
        return pd.DataFrame(
            columns=[
                "start_pos",
                "input_start_pos",
                "input_end_pos",
                "target_start_pos",
                "target_end_pos",
                "input_start_time",
                "input_end_time",
                "target_start_time",
                "target_end_time",
            ]
        )
    out = pd.DataFrame(
        {
            "start_pos": starts.to_numpy(),
            "input_start_pos": starts.to_numpy(),
            "input_end_pos": starts.to_numpy() + seq_len - 1,
            "target_start_pos": starts.to_numpy() + seq_len,
            "target_end_pos": starts.to_numpy() + total_len - 1,
        }
    )
    for col in ("input_start", "input_end", "target_start", "target_end"):
        pos_col = f"{col}_pos"
        time_col = f"{col}_time"
        out[time_col] = df["timestamp"].iloc[out[pos_col].to_numpy()].to_numpy()
    return out


def valve_response_diagnostics(pd: Any, df: Any, cfg: dict[str, Any]) -> Any:
    rows = []
    for item in cfg.get("valve_response_pairs", []):
        control = item.get("control")
        response = item.get("response")
        if control not in df.columns or response not in df.columns:
            rows.append({"name": item.get("name", ""), "control": control, "response": response, "status": "missing_column"})
            continue
        pair = df[[control, response]].dropna()
        if len(pair) < 100:
            rows.append({"name": item.get("name", ""), "control": control, "response": response, "status": "too_few_rows", "rows": len(pair)})
            continue
        pearson = pair[control].corr(pair[response], method="pearson")
        spearman = pair[control].rank().corr(pair[response].rank(), method="pearson")
        bins = min(20, max(4, int(math.sqrt(len(pair)))))
        binned = pair.copy()
        binned["_bin"] = pd.qcut(binned[control], q=bins, duplicates="drop")
        curve = binned.groupby("_bin", observed=True)[response].median()
        mono_ratio = float((curve.diff().dropna() >= 0).mean()) if len(curve) > 1 else math.nan
        rows.append(
            {
                "name": item.get("name", ""),
                "control": control,
                "response": response,
                "status": "ok",
                "rows": len(pair),
                "pearson": float(pearson) if pearson == pearson else math.nan,
                "spearman": float(spearman) if spearman == spearman else math.nan,
                "monotonic_increasing_ratio": mono_ratio,
                "note": "low_monotonic_or_nonlinear" if (mono_ratio == mono_ratio and mono_ratio < 0.7) else "",
            }
        )
    return pd.DataFrame(rows)


def write_dataset(pd: Any, df: Any, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        df.to_parquet(path, index=False)
    except Exception:
        fallback = path.with_suffix(".csv")
        df.to_csv(fallback, index=False, encoding="utf-8-sig")
        print(f"parquet write failed; csv fallback: {rel_path(fallback)}", flush=True)


def model_slug(model: str) -> str:
    return model.replace("子模型", "model")


def time_label(value: Any) -> str:
    return str(value).replace(":", "").replace("-", "").replace(" ", "T")


def downcast_numeric_frame(df: Any) -> Any:
    out = df.copy()
    for col in out.columns:
        if col == "timestamp" or str(out[col].dtype) == "bool":
            continue
        if str(out[col].dtype).startswith("float"):
            out[col] = out[col].astype("float32")
    return out


def cached_time_range(paths: dict[str, Path], point_codes: Iterable[str]) -> tuple[Any, Any]:
    duckdb, pd, _, _ = require_modules()
    con = duckdb.connect(":memory:")
    starts = []
    ends = []
    for code in unique_in_order(point_codes):
        path = cache_path(paths, code)
        if not path.exists():
            continue
        start_time, end_time = con.execute(
            f"SELECT min(timestamp), max(timestamp) FROM read_parquet('{sql_path(path)}')"
        ).fetchone()
        if start_time is not None and end_time is not None:
            starts.append(pd.Timestamp(start_time))
            ends.append(pd.Timestamp(end_time))
    if not starts or not ends:
        raise SystemExit("没有可用的逐点缓存，无法确定分块时间范围。")
    return min(starts), max(ends)


def iter_time_chunks(pd: Any, start: Any, end: Any, chunk_days: int) -> list[tuple[Any, Any]]:
    chunks = []
    cursor = pd.Timestamp(start)
    final = pd.Timestamp(end)
    step = pd.Timedelta(days=max(1, int(chunk_days)))
    while cursor < final:
        next_time = min(cursor + step, final)
        chunks.append((cursor, next_time))
        cursor = next_time
    return chunks


def global_quality_from_caches(
    pd: Any,
    paths: dict[str, Path],
    point_codes: list[str],
    cfg: dict[str, Any],
    *,
    start: Any | None = None,
    end: Any | None = None,
) -> Any:
    rows = []
    zero_abs_tolerance = float(cfg.get("zero_abs_tolerance", 1e-9))
    for idx, code in enumerate(point_codes, 1):
        path = cache_path(paths, code)
        if not path.exists():
            rows.append({"point_code": code, "status": "missing_column"})
            continue
        frame = read_cached_series(pd, paths, code, start=start, end=end)
        if frame is None or code not in frame.columns:
            rows.append({"point_code": code, "status": "missing_column"})
            continue
        s = frame[code]
        non_null = int(s.notna().sum())
        values = s.dropna()
        zero_ratio = float((values.abs() <= zero_abs_tolerance).mean()) if non_null else 1.0
        constant_ratio = float(values.value_counts(dropna=True, normalize=True).iloc[0]) if non_null else 1.0
        missing_rate = float(1 - non_null / len(s)) if len(s) else 1.0
        rows.append(
            {
                "point_code": code,
                "rows": int(len(s)),
                "non_null_rows": non_null,
                "missing_rate": missing_rate,
                "zero_ratio": zero_ratio,
                "constant_ratio": constant_ratio,
                "min": float(s.min(skipna=True)) if non_null else math.nan,
                "max": float(s.max(skipna=True)) if non_null else math.nan,
                "mean": float(s.mean(skipna=True)) if non_null else math.nan,
                "warn_missing": bool(missing_rate > float(cfg.get("missing_rate_warn", 0.3))) if len(s) else True,
                "warn_zero": bool(zero_ratio > float(cfg.get("zero_ratio_warn", 0.95))),
                "warn_constant": bool(constant_ratio > float(cfg.get("constant_ratio_warn", 0.95))),
            }
        )
        if idx % 25 == 0:
            print(f"  quality scanned {idx} point caches", flush=True)
    return pd.DataFrame(rows)


def build_model_dataset_chunked(
    cfg: dict[str, Any],
    specs: list[VariableSpec],
    model: str,
    *,
    freq_seconds: int,
    seq_len: int,
    pred_len: int,
    window_stride: int,
    start: str | None,
    end: str | None,
    chunk_days: int,
    write_chunk_wide: bool,
    force: bool,
) -> None:
    _, pd, _, _ = require_modules()
    paths = apply_cache_range(make_output_dirs(cfg, freq_seconds), freq_seconds, start, end)
    model_specs = [s for s in specs if s.model == model]
    model_codes = unique_in_order([s.point_code for s in model_specs])
    all_codes = unique_in_order(model_codes + sorted(derived_source_codes(cfg)))
    slug = model_slug(model)
    total_len = int(seq_len) + int(pred_len)
    if total_len <= 0:
        raise SystemExit("seq_len + pred_len 必须大于 0。")

    cache_start, cache_end = cached_time_range(paths, all_codes)
    full_start = pd.Timestamp(start) if start else cache_start
    full_end = pd.Timestamp(end) if end else cache_end + pd.Timedelta(seconds=freq_seconds)
    if full_start >= full_end:
        raise SystemExit(f"分块时间范围无效：{full_start} -> {full_end}")

    chunk_range_suffix = "" if not start and not end else f"_{time_label(full_start)}_{time_label(full_end)}"
    chunk_root = paths["datasets"] / "chunked" / f"{slug}_{freq_seconds}s_seq{seq_len}_pred{pred_len}{chunk_range_suffix}"
    frame_dir = chunk_root / "frames"
    wide_dir = chunk_root / "wide"
    window_dir = chunk_root / "windows"
    for path in (frame_dir, window_dir, wide_dir, chunk_root):
        path.mkdir(parents=True, exist_ok=True)

    print(f"chunked build for {model}: {len(model_codes)} model points", flush=True)
    print(f"range: {full_start} -> {full_end}; chunk_days={chunk_days}", flush=True)
    report_range_suffix = chunk_range_suffix
    q = global_quality_from_caches(pd, paths, model_codes, cfg, start=full_start, end=full_end)
    q_path = paths["reports"] / f"{slug}_quality_{freq_seconds}s{report_range_suffix}.csv"
    q.to_csv(q_path, index=False, encoding="utf-8-sig")
    low_info = low_information_inputs(pd, q, model_specs, cfg)
    low_info_path = paths["reports"] / f"low_information_inputs_{slug}_{freq_seconds}s{report_range_suffix}.csv"
    low_info.to_csv(low_info_path, index=False, encoding="utf-8-sig")
    bad_outputs = unusable_outputs(pd, q, model_specs, cfg)
    bad_outputs_path = paths["reports"] / f"unusable_outputs_{slug}_{freq_seconds}s{report_range_suffix}.csv"
    bad_outputs.to_csv(bad_outputs_path, index=False, encoding="utf-8-sig")
    stable_drop_cols = low_info["point_code"].tolist() if "point_code" in low_info.columns else []
    stable_bad_outputs = bad_outputs["point_code"].tolist() if "point_code" in bad_outputs.columns else []
    if stable_bad_outputs:
        print(f"{model}: drop unusable output columns from training targets: {len(stable_bad_outputs)}", flush=True)

    settings = cfg.get("long_common_low_anomaly", {})
    anomaly_context = int(max(float(settings.get("rolling_seconds", 1800)), float(settings.get("bridge_gap_seconds", 3600))))
    buffer_context = int(cfg.get("training_exclusion_buffer_seconds", 0))
    back_seconds = max(anomaly_context, buffer_context, 600)
    forward_seconds = max((total_len - 1) * int(freq_seconds), anomaly_context, buffer_context, 600)

    manifest_path = chunk_root / "manifest.json"
    manifest = load_manifest(manifest_path)
    manifest.setdefault("chunks", {})
    manifest_rows = []
    common_low_windows_all = []
    excluded_windows_all = []
    invalid_windows_all = []
    chunks = iter_time_chunks(pd, full_start, full_end, chunk_days)
    for idx, (core_start, core_end) in enumerate(chunks, 1):
        chunk_id = f"part{idx:04d}_{time_label(core_start)}_{time_label(core_end)}"
        frame_path = frame_dir / f"{chunk_id}.parquet"
        window_path = window_dir / f"{chunk_id}_windows.csv"
        wide_path = wide_dir / f"{chunk_id}_wide.parquet"
        param_id = params_hash(
            {
                "model": model,
                "freq_seconds": freq_seconds,
                "seq_len": seq_len,
                "pred_len": pred_len,
                "window_stride": window_stride,
                "variable_specs": [asdict(spec) for spec in model_specs],
                "core_start": str(core_start),
                "core_end": str(core_end),
                "drop_cols": stable_drop_cols,
                "drop_outputs": stable_bad_outputs,
                "target_forward_fill_limit": cfg_value(cfg, "target_forward_fill_limit", 5, model),
                "input_interpolate_limit": cfg_value(cfg, "input_interpolate_limit", 6, model),
                "input_forward_fill_limit": cfg_value(cfg, "input_forward_fill_limit", 30, model),
                "training_exclusion_buffer_seconds": cfg_value(cfg, "training_exclusion_buffer_seconds", 0, model),
                "write_chunk_wide": write_chunk_wide,
            }
        )
        item = manifest["chunks"].get(chunk_id)
        if (
            not force
            and item
            and item.get("param_id") == param_id
            and frame_path.exists()
            and window_path.exists()
        ):
            print(f"[{idx}/{len(chunks)}] skip {chunk_id}", flush=True)
            manifest_rows.append(item)
            continue

        load_start = max(full_start, core_start - pd.Timedelta(seconds=back_seconds))
        load_end = min(full_end, core_end + pd.Timedelta(seconds=forward_seconds))
        print(f"[{idx}/{len(chunks)}] build {chunk_id} load {load_start} -> {load_end}", flush=True)
        wide = merge_wide_table(
            pd,
            paths,
            all_codes,
            cfg,
            start=load_start,
            end=load_end,
            check_size=False,
        )
        wide, derived_notes = add_derived_features(pd, wide, cfg)
        wide = clean_model_frame(pd, wide, model_specs, cfg, model=model)
        if write_chunk_wide:
            write_dataset(pd, downcast_numeric_frame(wide), wide_path)

        outputs = usable_output_codes(model_specs, wide, stable_bad_outputs)
        if not outputs:
            raise SystemExit(f"{model}: 没有可用输出列，无法生成训练窗口。请检查 {rel_path(bad_outputs_path)}。")
        common_low_score, common_low_windows, _ = long_common_low_anomaly(
            pd,
            wide,
            cfg,
            model_codes,
            freq_seconds=freq_seconds,
        )
        training_mask = build_training_mask(
            pd,
            wide,
            cfg,
            model_codes,
            outputs,
            freq_seconds=freq_seconds,
            common_low_score=common_low_score,
            model=model,
        )
        training_frame = wide.copy()
        training_frame["valid_for_training"] = training_mask["valid_for_training"].to_numpy()
        if bool(cfg.get("drop_low_information_inputs", True)) and stable_drop_cols:
            drop_cols = [c for c in stable_drop_cols if c in training_frame.columns]
            training_frame = training_frame.drop(columns=drop_cols)
        if stable_bad_outputs:
            drop_outputs = [c for c in stable_bad_outputs if c in training_frame.columns]
            training_frame = training_frame.drop(columns=drop_outputs)
        training_frame = downcast_numeric_frame(training_frame)
        write_dataset(pd, training_frame, frame_path)

        window_index = build_window_index(
            pd,
            wide,
            training_mask["valid_for_training"],
            seq_len=seq_len,
            pred_len=pred_len,
            stride=window_stride,
            freq_seconds=freq_seconds,
        )
        if not window_index.empty:
            window_index = window_index[
                (window_index["input_start_time"] >= core_start) & (window_index["input_start_time"] < core_end)
            ].copy()
            window_index.insert(0, "chunk_id", chunk_id)
            window_index.insert(1, "frame_path", str(frame_path))
        window_index.to_csv(window_path, index=False, encoding="utf-8-sig")

        windows = detect_excluded_windows(pd, wide, cfg, model_codes)
        buffered_windows = boolean_windows(
            pd,
            training_mask["timestamp"],
            ~training_mask["valid_for_training"],
            "invalid_for_training_after_buffer",
        )
        for df_report, target in (
            (common_low_windows, common_low_windows_all),
            (windows, excluded_windows_all),
            (buffered_windows, invalid_windows_all),
        ):
            if not df_report.empty:
                df_report = df_report.copy()
                df_report.insert(0, "chunk_id", chunk_id)
                target.append(df_report)

        row = {
            "chunk_id": chunk_id,
            "param_id": param_id,
            "core_start": str(core_start),
            "core_end": str(core_end),
            "load_start": str(load_start),
            "load_end": str(load_end),
            "frame_path": str(frame_path),
            "window_path": str(window_path),
            "wide_path": str(wide_path) if write_chunk_wide else "",
            "rows": int(len(training_frame)),
            "core_rows": int(((training_frame["timestamp"] >= core_start) & (training_frame["timestamp"] < core_end)).sum()),
            "valid_rows": int(training_mask["valid_for_training"].sum()) if len(training_mask) else 0,
            "windows": int(len(window_index)),
            "dropped_low_information_inputs": int(len(stable_drop_cols)),
        }
        manifest["chunks"][chunk_id] = row
        manifest_rows.append(row)
        save_manifest(manifest_path, manifest)
        print(
            f"  chunk rows={row['rows']} core_rows={row['core_rows']} "
            f"valid_rows={row['valid_rows']} windows={row['windows']}",
            flush=True,
        )
        if derived_notes:
            notes_path = chunk_root / "derived_feature_notes.csv"
            pd.DataFrame(derived_notes).to_csv(notes_path, index=False, encoding="utf-8-sig")

    manifest_csv = chunk_root / "manifest.csv"
    pd.DataFrame(manifest_rows).to_csv(manifest_csv, index=False, encoding="utf-8-sig")
    if common_low_windows_all:
        pd.concat(common_low_windows_all, ignore_index=True).to_csv(
            paths["reports"] / f"common_low_anomaly_windows_{slug}_{freq_seconds}s{report_range_suffix}_chunked.csv",
            index=False,
            encoding="utf-8-sig",
        )
    if excluded_windows_all:
        pd.concat(excluded_windows_all, ignore_index=True).to_csv(
            paths["reports"] / f"excluded_time_windows_{slug}_{freq_seconds}s{report_range_suffix}_chunked.csv",
            index=False,
            encoding="utf-8-sig",
        )
    if invalid_windows_all:
        pd.concat(invalid_windows_all, ignore_index=True).to_csv(
            paths["reports"] / f"invalid_training_windows_{slug}_{freq_seconds}s{report_range_suffix}_chunked.csv",
            index=False,
            encoding="utf-8-sig",
        )
    print(f"chunked manifest: {rel_path(manifest_csv)}", flush=True)


def build_model_dataset(
    cfg: dict[str, Any],
    specs: list[VariableSpec],
    model: str,
    *,
    freq_seconds: int,
    seq_len: int,
    pred_len: int,
    window_stride: int,
    start: str | None = None,
    end: str | None = None,
) -> None:
    _, pd, _, _ = require_modules()
    paths = apply_cache_range(make_output_dirs(cfg, freq_seconds), freq_seconds, start, end)
    model_specs = [s for s in specs if s.model == model]
    model_codes = []
    seen = set()
    for spec in model_specs:
        if spec.point_code not in seen:
            seen.add(spec.point_code)
            model_codes.append(spec.point_code)
    print(f"build wide table for {model}: {len(model_codes)} model points", flush=True)
    wide = merge_wide_table(pd, paths, unique_in_order(model_codes + sorted(derived_source_codes(cfg))), cfg)
    wide, derived_notes = add_derived_features(pd, wide, cfg)
    wide = clean_model_frame(pd, wide, model_specs, cfg, model=model)

    model_slug = model.replace("子模型", "model")
    data_path = paths["datasets"] / f"{model_slug}_{freq_seconds}s.parquet"
    write_dataset(pd, wide, data_path)

    q = quality_summary(pd, wide, model_codes, cfg)
    q_path = paths["reports"] / f"{model_slug}_quality_{freq_seconds}s.csv"
    q.to_csv(q_path, index=False, encoding="utf-8-sig")

    low_info = low_information_inputs(pd, q, model_specs, cfg)
    low_info_path = paths["reports"] / f"low_information_inputs_{model_slug}_{freq_seconds}s.csv"
    low_info.to_csv(low_info_path, index=False, encoding="utf-8-sig")
    bad_outputs = unusable_outputs(pd, q, model_specs, cfg)
    bad_outputs_path = paths["reports"] / f"unusable_outputs_{model_slug}_{freq_seconds}s.csv"
    bad_outputs.to_csv(bad_outputs_path, index=False, encoding="utf-8-sig")
    stable_bad_outputs = bad_outputs["point_code"].tolist() if "point_code" in bad_outputs.columns else []

    outputs = usable_output_codes(model_specs, wide, stable_bad_outputs)
    if not outputs:
        raise SystemExit(f"{model}: 没有可用输出列，无法生成训练窗口。请检查 {rel_path(bad_outputs_path)}。")
    common_low_score, common_low_windows, common_low_candidates = long_common_low_anomaly(
        pd,
        wide,
        cfg,
        model_codes,
        freq_seconds=freq_seconds,
    )
    common_low_score_path = paths["reports"] / f"common_low_score_{model_slug}_{freq_seconds}s.parquet"
    write_dataset(pd, common_low_score, common_low_score_path)
    common_low_windows_path = paths["reports"] / f"common_low_anomaly_windows_{model_slug}_{freq_seconds}s.csv"
    common_low_windows.to_csv(common_low_windows_path, index=False, encoding="utf-8-sig")
    common_low_candidates_path = paths["reports"] / f"common_low_candidates_{model_slug}_{freq_seconds}s.csv"
    common_low_candidates.to_csv(common_low_candidates_path, index=False, encoding="utf-8-sig")

    training_mask = build_training_mask(
        pd,
        wide,
        cfg,
        model_codes,
        outputs,
        freq_seconds=freq_seconds,
        common_low_score=common_low_score,
        model=model,
    )
    training_frame = wide.copy()
    training_frame["valid_for_training"] = training_mask["valid_for_training"].to_numpy()

    if bool(cfg.get("drop_low_information_inputs", True)) and not low_info.empty:
        drop_cols = [c for c in low_info["point_code"].tolist() if c in training_frame.columns]
        training_frame = training_frame.drop(columns=drop_cols)
        print(f"{model}: drop low-information input columns from training frame: {len(drop_cols)}", flush=True)
    if stable_bad_outputs:
        drop_outputs = [c for c in stable_bad_outputs if c in training_frame.columns]
        training_frame = training_frame.drop(columns=drop_outputs)
        print(f"{model}: drop unusable output columns from training frame: {len(drop_outputs)}", flush=True)

    training_path = paths["datasets"] / f"{model_slug}_training_frame_{freq_seconds}s.parquet"
    write_dataset(pd, training_frame, training_path)

    mask_path = paths["reports"] / f"training_mask_{model_slug}_{freq_seconds}s.parquet"
    write_dataset(pd, training_mask, mask_path)

    windows = detect_excluded_windows(pd, wide, cfg, model_codes)
    w_path = paths["reports"] / f"excluded_time_windows_{model_slug}_{freq_seconds}s.csv"
    windows.to_csv(w_path, index=False, encoding="utf-8-sig")

    buffered_windows = boolean_windows(
        pd,
        training_mask["timestamp"],
        ~training_mask["valid_for_training"],
        "invalid_for_training_after_buffer",
    )
    bw_path = paths["reports"] / f"invalid_training_windows_{model_slug}_{freq_seconds}s.csv"
    buffered_windows.to_csv(bw_path, index=False, encoding="utf-8-sig")

    window_index = build_window_index(
        pd,
        wide,
        training_mask["valid_for_training"],
        seq_len=seq_len,
        pred_len=pred_len,
        stride=window_stride,
        freq_seconds=freq_seconds,
    )
    window_path = paths["reports"] / f"window_index_{model_slug}_{freq_seconds}s_seq{seq_len}_pred{pred_len}.csv"
    window_index.to_csv(window_path, index=False, encoding="utf-8-sig")

    valve = valve_response_diagnostics(pd, wide, cfg)
    v_path = paths["reports"] / f"valve_response_diagnostics_{model_slug}_{freq_seconds}s.csv"
    valve.to_csv(v_path, index=False, encoding="utf-8-sig")

    if derived_notes:
        n_path = paths["reports"] / f"derived_feature_notes_{model_slug}_{freq_seconds}s.csv"
        pd.DataFrame(derived_notes).to_csv(n_path, index=False, encoding="utf-8-sig")

    valid_rows = int(training_mask["valid_for_training"].sum()) if len(training_mask) else 0
    print(f"dataset: {rel_path(data_path)} rows={len(wide)} cols={len(wide.columns)}", flush=True)
    print(
        f"training frame: {rel_path(training_path)} valid_rows={valid_rows} "
        f"windows={len(window_index)} seq_len={seq_len} pred_len={pred_len}",
        flush=True,
    )
    print(
        f"reports: {rel_path(q_path)}, {rel_path(common_low_windows_path)}, "
        f"{rel_path(w_path)}, {rel_path(window_path)}, {rel_path(v_path)}",
        flush=True,
    )


def list_variables(args: argparse.Namespace) -> int:
    cfg = load_config(args.config)
    specs = load_variable_specs(cfg)
    point_root = Path(cfg["point_root"])
    available = available_point_codes(point_root)
    rows = []
    for spec in specs:
        rows.append(
            {
                **asdict(spec),
                "has_silver": spec.point_code in available,
            }
        )
    models = cfg.get("models", [])
    for model in models:
        subset = [r for r in rows if r["model"] == model]
        inputs = sum(1 for r in subset if r["io_type"] == "输入")
        outputs = sum(1 for r in subset if r["io_type"] == "输出")
        missing = sum(1 for r in subset if not r["has_silver"])
        print(f"{model}: inputs={inputs}, outputs={outputs}, missing_silver={missing}")
    if args.write:
        paths = make_output_dirs(cfg, int(cfg.get("default_freq_seconds", 60)))
        out = paths["reports"] / "model_variable_inventory.csv"
        fieldnames = list(rows[0].keys()) if rows else []
        with out.open("w", encoding="utf-8-sig", newline="") as f:
            writer = csv.DictWriter(f, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(rows)
        print(f"inventory: {rel_path(out)}")
    return 0


def inspect_points(args: argparse.Namespace) -> int:
    cfg = load_config(args.config)
    duckdb, _, _, _ = require_modules()
    specs = load_variable_specs(cfg)
    if args.codes:
        codes = [clean_text(x) for x in args.codes.split(",") if clean_text(x)]
    else:
        codes = [
            s.point_code
            for s in specs
            if (not args.model or s.model == args.model) and (not args.io_type or s.io_type == args.io_type)
        ]
    codes = unique_in_order(codes)
    point_root = Path(cfg["point_root"])
    con = duckdb.connect(":memory:")
    rows = []
    for idx, code in enumerate(codes, 1):
        src = point_path(point_root, code)
        print(f"[{idx}/{len(codes)}] inspect {code}", flush=True)
        if not src.exists():
            rows.append({"point_code": code, "status": "missing_source"})
            continue
        query = f"""
            SELECT
                count(*) AS raw_rows,
                count(value) FILTER (WHERE value IS NOT NULL AND isfinite(value)) AS valid_rows,
                min(timestamp) FILTER (WHERE value IS NOT NULL AND isfinite(value)) AS start_time,
                max(timestamp) FILTER (WHERE value IS NOT NULL AND isfinite(value)) AS end_time
            FROM read_parquet('{sql_path(src)}')
        """
        raw_rows, valid_rows, start_time, end_time = con.execute(query).fetchone()
        rows.append(
            {
                "point_code": code,
                "status": "ok",
                "raw_rows": raw_rows,
                "valid_rows": valid_rows,
                "valid_rate": float(valid_rows / raw_rows) if raw_rows else 0.0,
                "start_time": start_time or "",
                "end_time": end_time or "",
            }
        )
    if rows:
        fieldnames = list(rows[0].keys())
        for row in rows:
            print(
                f"{row.get('point_code')}: {row.get('status')} valid={row.get('valid_rows', '')} "
                f"{row.get('start_time', '')} -> {row.get('end_time', '')}"
            )
        if args.write:
            paths = make_output_dirs(cfg, int(cfg.get("default_freq_seconds", 60)))
            suffix = args.model or "selected"
            out = paths["reports"] / f"point_coverage_{suffix}.csv"
            with out.open("w", encoding="utf-8-sig", newline="") as f:
                writer = csv.DictWriter(f, fieldnames=fieldnames)
                writer.writeheader()
                writer.writerows(rows)
            print(f"coverage: {rel_path(out)}")
    return 0


def build(args: argparse.Namespace) -> int:
    cfg = load_config(args.config)
    freq_seconds = int(args.freq_seconds or cfg.get("default_freq_seconds", 60))
    seq_len = int(args.seq_len or cfg.get("window_seq_len", 1200))
    pred_len = int(args.pred_len or cfg.get("window_pred_len", 300))
    window_stride = int(args.window_stride or cfg.get("window_stride", 1))
    specs = load_variable_specs(cfg)
    models = set(cfg.get("models", [])) if args.all_models else {args.model}
    if not models:
        raise SystemExit("请指定 --model 或 --all-models")

    build_point_caches(
        cfg,
        specs,
        models,
        freq_seconds=freq_seconds,
        start=args.start,
        end=args.end,
        force=args.force,
        limit_points=args.limit_points,
    )
    if args.caches_only:
        return 0
    for model in sorted(models):
        if args.chunked:
            build_model_dataset_chunked(
                cfg,
                specs,
                model,
                freq_seconds=freq_seconds,
                seq_len=seq_len,
                pred_len=pred_len,
                window_stride=window_stride,
                start=args.start,
                end=args.end,
                chunk_days=args.chunk_days,
                write_chunk_wide=args.write_chunk_wide,
                force=args.force,
            )
        else:
            build_model_dataset(
                cfg,
                specs,
                model,
                freq_seconds=freq_seconds,
                seq_len=seq_len,
                pred_len=pred_len,
                window_stride=window_stride,
                start=args.start,
                end=args.end,
            )
    return 0


def load_resampled_points_for_audit(
    cfg: dict[str, Any],
    codes: list[str],
    *,
    freq_seconds: int,
    start: str | None,
    end: str | None,
) -> Any:
    duckdb, pd, _, _ = require_modules()
    point_root = Path(cfg["point_root"])
    con = duckdb.connect(":memory:")
    wide = None
    total = len(codes)
    for idx, code in enumerate(codes, 1):
        src = point_path(point_root, code)
        print(f"[{idx}/{total}] audit load {code}", flush=True)
        if not src.exists():
            print(f"  missing source: {code}", flush=True)
            continue
        filters = ["value IS NOT NULL", "isfinite(value)"]
        start_ts = sql_ts(start)
        end_ts = sql_ts(end)
        if start_ts:
            filters.append(f"timestamp >= TIMESTAMP '{start_ts}'")
        if end_ts:
            filters.append(f"timestamp < TIMESTAMP '{end_ts}'")
        query = f"""
            SELECT
                TIMESTAMP '1970-01-01'
                    + CAST(ceil(epoch(timestamp) / {freq_seconds}) * {freq_seconds} AS BIGINT) * INTERVAL '1 second'
                    AS timestamp,
                avg(value) AS "{code}"
            FROM read_parquet('{sql_path(src)}')
            WHERE {" AND ".join(filters)}
            GROUP BY 1
            ORDER BY 1
        """
        frame = con.execute(query).fetch_df()
        if frame.empty:
            continue
        frame["timestamp"] = pd.to_datetime(frame["timestamp"])
        wide = frame if wide is None else wide.merge(frame, on="timestamp", how="outer")
    if wide is None:
        return pd.DataFrame({"timestamp": pd.to_datetime([])})
    return wide.sort_values("timestamp", kind="stable").reset_index(drop=True)


def audit_anomalies(args: argparse.Namespace) -> int:
    cfg = load_config(args.config)
    _, pd, _, _ = require_modules()
    codes = [clean_text(x) for x in args.codes.split(",") if clean_text(x)]
    if not codes:
        raise SystemExit("audit-anomalies 需要 --codes，或修改 INTERACTIVE_AUDIT_POINT_CODES。")
    freq_seconds = int(args.freq_seconds)
    paths = make_output_dirs(cfg, freq_seconds)
    df = load_resampled_points_for_audit(
        cfg,
        codes,
        freq_seconds=freq_seconds,
        start=args.start,
        end=args.end,
    )
    score, windows, candidates = long_common_low_anomaly(
        pd,
        df,
        cfg,
        codes,
        freq_seconds=freq_seconds,
    )
    out_prefix = "audit_common_low"
    score_path = paths["reports"] / f"{out_prefix}_score_{freq_seconds}s.parquet"
    windows_path = paths["reports"] / f"{out_prefix}_windows_{freq_seconds}s.csv"
    candidates_path = paths["reports"] / f"{out_prefix}_candidates_{freq_seconds}s.csv"
    write_dataset(pd, score, score_path)
    windows.to_csv(windows_path, index=False, encoding="utf-8-sig")
    candidates.to_csv(candidates_path, index=False, encoding="utf-8-sig")
    print(f"audit rows={len(df)} candidates={len(candidates)} windows={len(windows)}", flush=True)
    if len(windows):
        print(windows.to_string(index=False), flush=True)
    print(f"audit reports: {rel_path(windows_path)}, {rel_path(candidates_path)}, {rel_path(score_path)}", flush=True)
    return 0


def interactive_argv() -> list[str]:
    action = INTERACTIVE_ACTION.strip()
    if action == "check-env":
        return ["check-env"]
    if action == "list-variables":
        argv = ["list-variables"]
        if INTERACTIVE_WRITE_REPORT:
            argv.append("--write")
        return argv
    if action == "inspect-points":
        argv = ["inspect-points"]
        if INTERACTIVE_POINT_CODES:
            argv += ["--codes", INTERACTIVE_POINT_CODES]
        else:
            argv += ["--model", INTERACTIVE_MODEL]
            if INTERACTIVE_IO_TYPE:
                argv += ["--io-type", INTERACTIVE_IO_TYPE]
        if INTERACTIVE_WRITE_REPORT:
            argv.append("--write")
        return argv
    if action == "audit-anomalies":
        argv = ["audit-anomalies", "--codes", INTERACTIVE_AUDIT_POINT_CODES]
        argv += ["--freq-seconds", str(INTERACTIVE_AUDIT_FREQ_SECONDS)]
        if INTERACTIVE_START:
            argv += ["--start", INTERACTIVE_START]
        if INTERACTIVE_END:
            argv += ["--end", INTERACTIVE_END]
        return argv

    argv = ["build"]
    if INTERACTIVE_ALL_MODELS:
        argv.append("--all-models")
    else:
        argv += ["--model", INTERACTIVE_MODEL]
    argv += ["--freq-seconds", str(INTERACTIVE_FREQ_SECONDS)]
    if INTERACTIVE_START:
        argv += ["--start", INTERACTIVE_START]
    if INTERACTIVE_END:
        argv += ["--end", INTERACTIVE_END]
    if INTERACTIVE_FORCE_REBUILD_CACHES:
        argv.append("--force")
    if INTERACTIVE_LIMIT_POINTS:
        argv += ["--limit-points", str(INTERACTIVE_LIMIT_POINTS)]
    if INTERACTIVE_CACHES_ONLY:
        argv.append("--caches-only")
    if INTERACTIVE_CHUNKED:
        argv.append("--chunked")
    if INTERACTIVE_CHUNK_DAYS:
        argv += ["--chunk-days", str(INTERACTIVE_CHUNK_DAYS)]
    if INTERACTIVE_WRITE_CHUNK_WIDE:
        argv.append("--write-chunk-wide")
    argv += ["--seq-len", str(INTERACTIVE_SEQ_LEN)]
    argv += ["--pred-len", str(INTERACTIVE_PRED_LEN)]
    argv += ["--window-stride", str(INTERACTIVE_WINDOW_STRIDE)]
    return argv


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build model-ready datasets from silver point parquet files.")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG, help="JSON config path.")
    sub = parser.add_subparsers(dest="command", required=True)

    p_env = sub.add_parser("check-env", help="Check Python dependencies and configured paths.")
    p_env.set_defaults(func=check_env)

    p_list = sub.add_parser("list-variables", help="List model variable counts and missing silver points.")
    p_list.add_argument("--write", action="store_true", help="Write variable inventory CSV.")
    p_list.set_defaults(func=list_variables)

    p_inspect = sub.add_parser("inspect-points", help="Inspect raw silver coverage for selected points.")
    p_inspect.add_argument("--codes", default="", help="Comma-separated point codes. Overrides --model/--io-type.")
    p_inspect.add_argument("--model", default="", help="Model name to inspect, e.g. 子模型2.")
    p_inspect.add_argument("--io-type", default="", help="输入 or 输出.")
    p_inspect.add_argument("--write", action="store_true", help="Write coverage CSV.")
    p_inspect.set_defaults(func=inspect_points)

    p_audit = sub.add_parser("audit-anomalies", help="Audit long common low anomalies for selected points.")
    p_audit.add_argument("--codes", required=True, help="Comma-separated point codes.")
    p_audit.add_argument("--freq-seconds", type=int, default=300, help="Audit resample interval in seconds.")
    p_audit.add_argument("--start", default=None, help="Inclusive start timestamp.")
    p_audit.add_argument("--end", default=None, help="Exclusive end timestamp.")
    p_audit.set_defaults(func=audit_anomalies)

    p_build = sub.add_parser("build", help="Build point caches and model datasets.")
    p_build.add_argument("--model", default="子模型2", help="Model name, e.g. 子模型1/子模型2/子模型3.")
    p_build.add_argument("--all-models", action="store_true", help="Build all configured models.")
    p_build.add_argument("--freq-seconds", type=int, default=None, help="Resample interval in seconds.")
    p_build.add_argument("--start", default=None, help="Inclusive start timestamp, e.g. 2025-05-11.")
    p_build.add_argument("--end", default=None, help="Exclusive end timestamp, e.g. 2025-05-12.")
    p_build.add_argument("--force", action="store_true", help="Rebuild point caches even if checkpoint is valid.")
    p_build.add_argument("--limit-points", type=int, default=0, help="Only process first N source points for smoke tests.")
    p_build.add_argument("--caches-only", action="store_true", help="Only build per-point resampled caches.")
    p_build.add_argument("--chunked", action="store_true", help="Build chunked training frames instead of one full wide table.")
    p_build.add_argument("--chunk-days", type=int, default=1, help="Core days per chunk when --chunked is enabled.")
    p_build.add_argument("--write-chunk-wide", action="store_true", help="Also write cleaned wide table chunks.")
    p_build.add_argument("--seq-len", type=int, default=None, help="Sliding-window input length in rows.")
    p_build.add_argument("--pred-len", type=int, default=None, help="Sliding-window prediction length in rows.")
    p_build.add_argument("--window-stride", type=int, default=None, help="Sliding-window stride in rows.")
    p_build.set_defaults(func=build)
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    if argv is None and len(sys.argv) == 1:
        argv = interactive_argv()
        print("VSCode interactive argv:", " ".join(argv), flush=True)
    args = parse_args(argv)
    return int(args.func(args))


if __name__ == "__main__":
    raise SystemExit(main())
